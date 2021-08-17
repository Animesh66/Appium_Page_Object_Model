import openpyxl


def get_data(sheetName):
    workbook = openpyxl.load_workbook("/Users/animeshmukherjee/PycharmProjects/pythonProject/Appium_Page_Object_Model/Excel/appium_data.xlsx")
    sheet = workbook[sheetName]
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    mainList = []

    for i in range(2,totalrows+1):
        dataList = []
        for j in range(1, totalcols+1):
            data = sheet.cell(row=i,column=j).value
            dataList.insert(j,data)
        mainList.insert(i,dataList)
        print(mainList)
    return mainList